"""Excel/CSV data import endpoints (validate + commit 两步式).

每个 section 对应一类业务数据。validate 端点解析文件、跑校验、返回错误/警告/预览；
commit 端点重新解析同一份文件、跑校验、校验通过则落库。

支持的 section（按 plan_id 绑定）：
- production-tasks  → biz_production_task
- material-supply   → biz_material_supply
- inventory         → biz_inventory_snapshot
- wip               → biz_wip_buffer_snapshot

主数据型 section（bop / equipment-config / staffing / changeover / op-transition /
calendar / equipment-params）当前返回 501 —— 主数据按设计应从主数据平台 ETL 同步，
不走手工导入路径。"""

import csv
import io
from datetime import datetime
from decimal import Decimal, InvalidOperation
from typing import Any, Callable

from fastapi import APIRouter, Depends, File, Form, HTTPException, UploadFile
from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.api.deps import get_db
from app.models.biz import (
    InventorySnapshot,
    MaterialSupply,
    ProductionTask,
    WIPBufferSnapshot,
    WorkOrder,
)
from app.models.md import ProductionLine, Stage, Warehouse, WIPBuffer
from app.models.sim import SimulationPlan
from app.schemas.sim import (
    ImportCommitResult,
    ImportIssue,
    ImportValidationResult,
)

router = APIRouter(prefix="/imports", tags=["Data Import"])


# ---------------------------------------------------------------------------
# File parsing
# ---------------------------------------------------------------------------
_MAX_FILE_MB = 10
_MAX_PREVIEW = 5


def _read_upload(file: UploadFile) -> tuple[list[str], list[list[str]]]:
    """统一把 .xlsx / .xls / .csv 解析成 (header, rows)，全部字符串。"""
    name = (file.filename or "").lower()
    raw = file.file.read()
    if len(raw) > _MAX_FILE_MB * 1024 * 1024:
        raise HTTPException(400, f"文件超过 {_MAX_FILE_MB}MB 上限")
    if not raw:
        raise HTTPException(400, "文件为空")

    if name.endswith(".csv"):
        # 默认 UTF-8，BOM 自动剥离
        text = raw.decode("utf-8-sig", errors="replace")
        reader = csv.reader(io.StringIO(text))
        all_rows = [list(row) for row in reader if row]
    elif name.endswith((".xlsx", ".xls")):
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb.active
        all_rows = []
        for row in ws.iter_rows(values_only=True):
            if all(cell is None or str(cell).strip() == "" for cell in row):
                continue
            all_rows.append(["" if c is None else str(c).strip() for c in row])
        wb.close()
    else:
        raise HTTPException(400, "仅支持 .xlsx / .xls / .csv")

    if not all_rows:
        raise HTTPException(400, "未读取到有效数据")
    header = [h.strip() for h in all_rows[0]]
    body = all_rows[1:]
    # 把每行 pad 到 header 长度
    body = [row + [""] * (len(header) - len(row)) for row in body]
    body = [row[:len(header)] for row in body]
    return header, body


def _to_dicts(header: list[str], rows: list[list[str]]) -> list[dict[str, str]]:
    return [dict(zip(header, r)) for r in rows]


def _require_columns(header: list[str], required: list[str]) -> list[ImportIssue]:
    """检查 header 是否包含所有 required 列。返回 row=1 级错误。"""
    issues: list[ImportIssue] = []
    for col in required:
        if col not in header:
            issues.append(ImportIssue(row=1, field=col, message=f"缺少必填列：{col}"))
    return issues


def _parse_decimal(v: str) -> Decimal | None:
    s = (v or "").strip().replace(",", "")
    if not s:
        return None
    try:
        return Decimal(s)
    except InvalidOperation:
        return None


def _parse_int(v: str) -> int | None:
    d = _parse_decimal(v)
    if d is None:
        return None
    try:
        return int(d)
    except (ValueError, InvalidOperation):
        return None


def _parse_datetime(v: str) -> datetime | None:
    s = (v or "").strip()
    if not s:
        return None
    for fmt in (
        "%Y-%m-%d %H:%M:%S",
        "%Y-%m-%d %H:%M",
        "%Y-%m-%d",
        "%Y/%m/%d %H:%M:%S",
        "%Y/%m/%d %H:%M",
        "%Y/%m/%d",
    ):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Section validators / committers
# ---------------------------------------------------------------------------
def _plan_snapshotted(db: Session, plan: SimulationPlan) -> bool:
    return bool(
        plan.base_data_version and plan.base_data_version.startswith("snapshot:")
    )


def _line_lookup_for_plan(db: Session, plan: SimulationPlan) -> dict[str, ProductionLine]:
    """build line_code / line_name → ProductionLine 查找表。

    已快照方案：只取本方案 scoped 行（导入的 task 直接引用 scoped id，引擎才查得到）。
    未快照方案：维持原行为（同工厂全局行）。
    """
    q = db.query(ProductionLine).filter(ProductionLine.status == "ACTIVE")
    if _plan_snapshotted(db, plan):
        q = q.filter(ProductionLine.plan_id == plan.plan_id)
    else:
        q = q.join(Stage, Stage.stage_id == ProductionLine.stage_id).filter(
            Stage.factory_id == plan.factory_id, ProductionLine.plan_id.is_(None)
        )
    out: dict[str, ProductionLine] = {}
    for ln in q.all():
        out[ln.line_code] = ln
        out[ln.line_name] = ln
    return out


def _warehouse_lookup_for_plan(db: Session, plan: SimulationPlan) -> dict[str, Warehouse]:
    q = db.query(Warehouse)
    if _plan_snapshotted(db, plan):
        q = q.filter(Warehouse.plan_id == plan.plan_id)
    else:
        q = q.filter(Warehouse.factory_id == plan.factory_id, Warehouse.plan_id.is_(None))
    out: dict[str, Warehouse] = {}
    for w in q.all():
        out[w.warehouse_code] = w
        out[w.warehouse_name] = w
    return out


def _wip_lookup_for_plan(db: Session, plan: SimulationPlan) -> dict[str, WIPBuffer]:
    q = db.query(WIPBuffer).filter(WIPBuffer.status == "ACTIVE")
    if _plan_snapshotted(db, plan):
        q = q.filter(WIPBuffer.plan_id == plan.plan_id)
    else:
        q = (
            q.join(ProductionLine, ProductionLine.line_id == WIPBuffer.line_id)
            .join(Stage, Stage.stage_id == ProductionLine.stage_id)
            .filter(Stage.factory_id == plan.factory_id, WIPBuffer.plan_id.is_(None))
        )
    out: dict[str, WIPBuffer] = {}
    for w in q.all():
        out[w.wip_code] = w
        out[w.wip_name] = w
    return out


def _wo_lookup_for_plan(db: Session, plan: SimulationPlan) -> dict[str, WorkOrder]:
    """wo_no → WorkOrder。已快照方案取本方案 scoped 工单；未快照取全局工单。

    WO 当主数据看（seed 提供、建方案随快照克隆），不支持手工导入。
    """
    q = db.query(WorkOrder)
    if _plan_snapshotted(db, plan):
        q = q.filter(WorkOrder.plan_id == plan.plan_id)
    else:
        q = q.filter(WorkOrder.plan_id.is_(None))
    return {w.wo_no: w for w in q.all()}


# ── production-tasks ────────────────────────────────────────────────────────
PT_COLS = ["工单号", "产线", "产品型号", "计划产量"]


def _validate_production_tasks(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[list[ImportIssue], list[ImportIssue]]:
    errors: list[ImportIssue] = list(_require_columns(header, PT_COLS))
    warnings: list[ImportIssue] = []
    if errors:
        return errors, warnings
    line_lookup = _line_lookup_for_plan(db, plan)
    wo_lookup = _wo_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    seen_wo_line: set[tuple[str, str]] = set()
    for i, d in enumerate(dicts, start=2):
        wo_no = d["工单号"].strip()
        line_key = d["产线"].strip()
        product = d["产品型号"].strip()
        qty_raw = d["计划产量"].strip()
        if not wo_no:
            errors.append(ImportIssue(row=i, field="工单号", message="必填字段为空"))
        elif wo_no not in wo_lookup:
            errors.append(ImportIssue(
                row=i, field="工单号",
                message=f"工单号 {wo_no!r} 在本方案中不存在（工单由 seed/ERP 提供，"
                        f"不支持手工导入；请确认工单号或先准备好工单数据）",
            ))
        if not product:
            errors.append(ImportIssue(row=i, field="产品型号", message="必填字段为空"))
        ln = line_lookup.get(line_key)
        if not ln:
            errors.append(ImportIssue(
                row=i, field="产线",
                message=f"找不到产线（应为 line_code 或 line_name）: {line_key!r}",
            ))
        qty = _parse_int(qty_raw)
        if qty is None or qty <= 0:
            errors.append(ImportIssue(row=i, field="计划产量", message=f"无效数量：{qty_raw!r}"))
        key = (wo_no, line_key)
        if key in seen_wo_line:
            warnings.append(ImportIssue(
                row=i, field="工单号",
                message=f"工单 {wo_no} × 产线 {line_key} 在本表内重复",
            ))
        seen_wo_line.add(key)
    return errors, warnings


def _commit_production_tasks(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[int, int]:
    line_lookup = _line_lookup_for_plan(db, plan)
    wo_lookup = _wo_lookup_for_plan(db, plan)  # wo_no → 已存在 WorkOrder（不自动建）
    dicts = _to_dicts(header, rows)
    # 覆盖语义：导入即用本表全量替换该方案的生产任务，而非追加。
    # 与 commit_import 的单次 db.commit() 同事务，删+插原子（校验已先于此拦 error）。
    db.query(ProductionTask).filter(
        ProductionTask.plan_id == plan.plan_id
    ).delete(synchronize_session=False)
    inserted = 0
    skipped = 0

    for idx, d in enumerate(dicts, start=1):
        line_key = d["产线"].strip()
        ln = line_lookup.get(line_key)
        qty = _parse_int(d["计划产量"])
        product = d["产品型号"].strip()
        wo_no = d.get("工单号", "").strip()
        wo = wo_lookup.get(wo_no) if wo_no else None
        # 工单号填了却查不到 → 与校验一致地跳过（_validate 已会报 error 拦截 commit）
        if not ln or qty is None or qty <= 0 or not product or (wo_no and wo is None):
            skipped += 1
            continue
        db.add(ProductionTask(
            plan_id=plan.plan_id,
            wo_id=wo.wo_id if wo else None,
            stage_id=ln.stage_id,
            line_id=ln.line_id,
            product_code=product,
            plan_quantity=qty,
            production_sequence=idx,
            data_source="MANUAL_IMPORT",
            source_system="UPLOAD",
            sync_time=datetime.utcnow(),
        ))
        inserted += 1
    return inserted, skipped


# ── material-supply ─────────────────────────────────────────────────────────
MS_COLS = ["物料编码", "物料名称", "供应数量", "到货时间", "仓库"]


def _validate_material_supply(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[list[ImportIssue], list[ImportIssue]]:
    errors: list[ImportIssue] = list(_require_columns(header, MS_COLS))
    warnings: list[ImportIssue] = []
    if errors:
        return errors, warnings
    wh_lookup = _warehouse_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    for i, d in enumerate(dicts, start=2):
        if not d["物料编码"].strip():
            errors.append(ImportIssue(row=i, field="物料编码", message="必填字段为空"))
        qty = _parse_decimal(d["供应数量"])
        if qty is None or qty <= 0:
            errors.append(ImportIssue(
                row=i, field="供应数量", message=f"无效数量：{d['供应数量']!r}"
            ))
        # 到货时间支持两种语义：1) datetime 字段→相对 plan 起点的 sim_hour ；2) 直接小时数
        # 优先按小时数解析
        hours = _parse_decimal(d["到货时间"])
        if hours is None:
            dt = _parse_datetime(d["到货时间"])
            if dt is None:
                errors.append(ImportIssue(
                    row=i, field="到货时间",
                    message=f"到货时间需为小时数或日期时间：{d['到货时间']!r}",
                ))
        wh = wh_lookup.get(d["仓库"].strip())
        if not wh:
            errors.append(ImportIssue(
                row=i, field="仓库",
                message=f"找不到仓库（应为 warehouse_code 或 warehouse_name）: {d['仓库']!r}",
            ))
    return errors, warnings


def _commit_material_supply(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[int, int]:
    wh_lookup = _warehouse_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    inserted = 0
    skipped = 0
    for d in dicts:
        wh = wh_lookup.get(d["仓库"].strip())
        qty = _parse_decimal(d["供应数量"])
        # 解析到货时间：先小时数，再 datetime
        hours = _parse_decimal(d["到货时间"])
        if hours is None:
            dt = _parse_datetime(d["到货时间"])
            if dt is None:
                skipped += 1
                continue
            # 把 datetime 当作小时数 0 入库（更复杂的语义留给 ETL）
            hours = Decimal("0")
        if not wh or qty is None or qty <= 0:
            skipped += 1
            continue
        db.add(MaterialSupply(
            plan_id=plan.plan_id,
            material_code=d["物料编码"].strip(),
            material_name=d.get("物料名称", "").strip() or None,
            supply_quantity=qty,
            arrival_sim_hour=hours,
            target_warehouse_id=wh.warehouse_id,
            data_source="MANUAL_IMPORT",
            sync_time=datetime.utcnow(),
        ))
        inserted += 1
    return inserted, skipped


# ── inventory ───────────────────────────────────────────────────────────────
INV_COLS = ["仓库", "物料编码", "库存总量", "可用量", "快照时间"]


def _validate_inventory(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[list[ImportIssue], list[ImportIssue]]:
    errors: list[ImportIssue] = list(_require_columns(header, INV_COLS))
    warnings: list[ImportIssue] = []
    if errors:
        return errors, warnings
    wh_lookup = _warehouse_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    for i, d in enumerate(dicts, start=2):
        if not wh_lookup.get(d["仓库"].strip()):
            errors.append(ImportIssue(row=i, field="仓库", message=f"找不到仓库：{d['仓库']!r}"))
        if not d["物料编码"].strip():
            errors.append(ImportIssue(row=i, field="物料编码", message="必填字段为空"))
        total = _parse_decimal(d["库存总量"])
        avail = _parse_decimal(d["可用量"])
        if total is None or total < 0:
            errors.append(ImportIssue(row=i, field="库存总量", message=f"无效：{d['库存总量']!r}"))
        if avail is None or avail < 0:
            errors.append(ImportIssue(row=i, field="可用量", message=f"无效：{d['可用量']!r}"))
        if total is not None and avail is not None and avail > total:
            warnings.append(ImportIssue(
                row=i, field="可用量", message="可用量大于总量，将自动截断到总量",
            ))
        if _parse_datetime(d["快照时间"]) is None:
            errors.append(ImportIssue(row=i, field="快照时间", message=f"日期格式无效：{d['快照时间']!r}"))
    return errors, warnings


def _commit_inventory(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[int, int]:
    wh_lookup = _warehouse_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    inserted = 0
    skipped = 0
    for d in dicts:
        wh = wh_lookup.get(d["仓库"].strip())
        total = _parse_decimal(d["库存总量"])
        avail = _parse_decimal(d["可用量"])
        ts = _parse_datetime(d["快照时间"])
        if not wh or total is None or avail is None or ts is None:
            skipped += 1
            continue
        if avail > total:
            avail = total
        db.add(InventorySnapshot(
            plan_id=plan.plan_id,
            warehouse_id=wh.warehouse_id,
            material_code=d["物料编码"].strip(),
            total_quantity=total,
            available_quantity=avail,
            snapshot_time=ts,
            data_source="MANUAL_IMPORT",
        ))
        inserted += 1
    return inserted, skipped


# ── wip ─────────────────────────────────────────────────────────────────────
WIP_COLS = ["线边仓", "物料编码", "当前数量", "占用体积", "快照时间"]


def _validate_wip(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[list[ImportIssue], list[ImportIssue]]:
    errors: list[ImportIssue] = list(_require_columns(header, WIP_COLS))
    warnings: list[ImportIssue] = []
    if errors:
        return errors, warnings
    wip_lookup = _wip_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    for i, d in enumerate(dicts, start=2):
        wip = wip_lookup.get(d["线边仓"].strip())
        if not wip:
            errors.append(ImportIssue(row=i, field="线边仓", message=f"找不到线边仓：{d['线边仓']!r}"))
        if not d["物料编码"].strip():
            errors.append(ImportIssue(row=i, field="物料编码", message="必填字段为空"))
        qty = _parse_decimal(d["当前数量"])
        vol = _parse_decimal(d["占用体积"])
        if qty is None or qty < 0:
            errors.append(ImportIssue(row=i, field="当前数量", message=f"无效：{d['当前数量']!r}"))
        if vol is None or vol < 0:
            errors.append(ImportIssue(row=i, field="占用体积", message=f"无效：{d['占用体积']!r}"))
        if wip and vol is not None and vol > wip.capacity_volume:
            warnings.append(ImportIssue(
                row=i, field="占用体积",
                message=f"占用体积 {vol} 超过线边仓容量 {wip.capacity_volume}",
            ))
        if _parse_datetime(d["快照时间"]) is None:
            errors.append(ImportIssue(row=i, field="快照时间", message=f"日期格式无效：{d['快照时间']!r}"))
    return errors, warnings


def _commit_wip(
    db: Session, plan: SimulationPlan, header: list[str], rows: list[list[str]]
) -> tuple[int, int]:
    wip_lookup = _wip_lookup_for_plan(db, plan)
    dicts = _to_dicts(header, rows)
    inserted = 0
    skipped = 0
    for d in dicts:
        wip = wip_lookup.get(d["线边仓"].strip())
        qty = _parse_decimal(d["当前数量"])
        vol = _parse_decimal(d["占用体积"])
        ts = _parse_datetime(d["快照时间"])
        if not wip or qty is None or vol is None or ts is None:
            skipped += 1
            continue
        db.add(WIPBufferSnapshot(
            plan_id=plan.plan_id,
            wip_id=wip.wip_id,
            material_code=d["物料编码"].strip(),
            current_quantity=qty,
            current_volume=vol,
            snapshot_time=ts,
            data_source="MANUAL_IMPORT",
        ))
        inserted += 1
    return inserted, skipped


# ---------------------------------------------------------------------------
# Section registry
# ---------------------------------------------------------------------------
ValidatorFn = Callable[[Session, SimulationPlan, list[str], list[list[str]]],
                       tuple[list[ImportIssue], list[ImportIssue]]]
CommitterFn = Callable[[Session, SimulationPlan, list[str], list[list[str]]],
                       tuple[int, int]]

_PLAN_SCOPED_SECTIONS: dict[str, tuple[ValidatorFn, CommitterFn, list[str]]] = {
    "production-tasks": (_validate_production_tasks, _commit_production_tasks, PT_COLS),
    "material-supply": (_validate_material_supply, _commit_material_supply, MS_COLS),
    "inventory": (_validate_inventory, _commit_inventory, INV_COLS),
    "wip": (_validate_wip, _commit_wip, WIP_COLS),
}

# 主数据型 section：手工导入暂不支持，前端给提示
_MD_SECTIONS = {
    "bop": "BOP",
    "equipment-config": "产线设备配置",
    "staffing": "人员配置",
    "changeover": "换线时间",
    "op-transition": "工序间接续",
    "calendar": "工作日历 & 班次",
    "equipment-params": "设备故障参数",
}


def _resolve_section(section_id: str) -> tuple[ValidatorFn, CommitterFn, list[str]]:
    if section_id in _PLAN_SCOPED_SECTIONS:
        return _PLAN_SCOPED_SECTIONS[section_id]
    if section_id in _MD_SECTIONS:
        raise HTTPException(
            501,
            f"主数据型 section '{section_id}' ({_MD_SECTIONS[section_id]}) 当前不支持手工导入；"
            f"请通过主数据平台 ETL 同步",
        )
    raise HTTPException(404, f"未知 section: {section_id}")


# ---------------------------------------------------------------------------
# Endpoints
# ---------------------------------------------------------------------------
@router.post("/{section_id}:validate", response_model=ImportValidationResult)
def validate_import(
    section_id: str,
    plan_id: str = Form(...),
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
):
    """上传 .xlsx/.xls/.csv，按 section 校验数据并返回错误/警告/预览（不落库）。"""
    plan = db.query(SimulationPlan).filter(SimulationPlan.plan_id == plan_id).first()
    if not plan:
        raise HTTPException(404, "Plan not found")
    validator, _committer, required_cols = _resolve_section(section_id)

    header, rows = _read_upload(file)
    errors, warnings = validator(db, plan, header, rows)
    valid_rows = max(0, len(rows) - len({i.row for i in errors if i.row >= 2}))

    return ImportValidationResult(
        section_id=section_id,
        valid=len(errors) == 0,
        total_rows=len(rows),
        valid_rows=valid_rows,
        errors=errors,
        warnings=warnings,
        columns=header,
        preview_rows=[r[:len(header)] for r in rows[:_MAX_PREVIEW]],
    )


@router.post("/{section_id}:commit", response_model=ImportCommitResult)
def commit_import(
    section_id: str,
    plan_id: str = Form(...),
    file: UploadFile = File(...),
    ignore_warnings: bool = Form(True),
    db: Session = Depends(get_db),
):
    """重新解析同一份文件并落库。校验有 error 一律拒绝；warning 可由 `ignore_warnings` 控制。"""
    plan = db.query(SimulationPlan).filter(SimulationPlan.plan_id == plan_id).first()
    if not plan:
        raise HTTPException(404, "Plan not found")
    if plan.status not in ("DRAFT", "READY"):
        raise HTTPException(400, f"Plan 处于 {plan.status} 状态，不可导入数据")
    validator, committer, _required_cols = _resolve_section(section_id)

    header, rows = _read_upload(file)
    errors, warnings = validator(db, plan, header, rows)
    if errors:
        raise HTTPException(
            422,
            detail={
                "message": f"导入失败：{len(errors)} 条错误",
                "errors": [e.model_dump() for e in errors[:20]],
            },
        )
    if warnings and not ignore_warnings:
        raise HTTPException(
            422,
            detail={
                "message": f"存在 {len(warnings)} 条警告，未确认忽略，已拒绝导入",
                "warnings": [w.model_dump() for w in warnings[:20]],
            },
        )

    inserted, skipped = committer(db, plan, header, rows)
    # 导入改了输入数据 → 已就绪方案退回草稿，须重新过「保存并就绪」门
    if plan.status == "READY":
        plan.status = "DRAFT"
    db.commit()
    return ImportCommitResult(
        section_id=section_id,
        inserted=inserted,
        skipped=skipped,
        plan_id=plan_id,
        message=f"导入完成：成功 {inserted} 条 / 跳过 {skipped} 条",
    )
